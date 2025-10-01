from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook

from inventory.models import Category, Product, Warehouse, Supplier, Stock, StockProduct, Inventory
from reports.exporters import ExcelReportExporter
from reports.services.inventory import InventoryValuationReportBuilder


User = get_user_model()


class InventoryValuationReportTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            email='owner@example.com',
            password='pass12345',
            name='Owner'
        )
        self.category = Category.objects.create(name='Groceries')
        self.product = Product.objects.create(
            name='Rice Bag 25kg',
            sku='RICE-25KG',
            category=self.category
        )
        self.warehouse = Warehouse.objects.create(name='Central Warehouse', location='Industrial Estate')
        self.supplier = Supplier.objects.create(name='Agro Imports Ltd.')
        today = timezone.now().date()
        self.batch = Stock.objects.create(
            warehouse=self.warehouse,
            arrival_date=today,
            description='Goods receipt GRN-VAL-001'
        )
        self.stock = StockProduct.objects.create(
            stock=self.batch,
            product=self.product,
            supplier=self.supplier,
            expiry_date=today + timedelta(days=180),
            quantity=100,
            unit_cost=Decimal('200.00'),
            unit_tax_rate=Decimal('12.50'),
            unit_additional_cost=Decimal('15.00'),
            description='Lot STK-VAL-001'
        )
        Inventory.objects.create(
            product=self.product,
            warehouse=self.warehouse,
            stock=self.stock,
            quantity=80
        )

    def test_builder_outputs_expected_summary(self):
        builder = InventoryValuationReportBuilder(user=self.user)
        report = builder.build(filters={})

        self.assertEqual(report['summary']['total_rows'], 1)
        self.assertEqual(report['summary']['total_quantity'], 80)
        # Landed cost per unit: 200 + (200 * 12.5%) + 15 = 240
        self.assertEqual(report['summary']['inventory_value'], Decimal('19200.00'))
        self.assertEqual(report['summary']['total_tax_amount'], Decimal('2000.00'))

    def test_excel_exporter_generates_workbook(self):
        builder = InventoryValuationReportBuilder(user=self.user)
        report = builder.build(filters={})

        exporter = ExcelReportExporter()
        workbook_bytes = exporter.export(report)

        wb = load_workbook(filename=BytesIO(workbook_bytes))
        self.assertIn('Summary', wb.sheetnames)
        self.assertIn('Inventory Detail', wb.sheetnames)

        detail_sheet = wb['Inventory Detail']
        self.assertEqual(detail_sheet.max_row, 2)  # header + 1 row
        self.assertEqual(detail_sheet['E2'].value, 80)
